import re
import os
import pandas as pd
import openpyxl
from rapidfuzz import process as rfprocess


LOT_PATTERN = re.compile(r'\blot\s*[\d\w]+', re.IGNORECASE)

CANONICAL_CANDIDATES = {
    "description": ["description", "item description", "item name", "work item",
                     "particulars", "scope", "activity", "goods/services", "goods", "services", "item"],
    "unit":        ["unit", "uom", "unit of measure", "measure"],
    # "no." and "nos" intentionally excluded — they match serial-number columns, not quantities
    "quantity":    ["qty", "quantity", "number of units", "count", "volume"],
    "unit_price":  ["unit price", "rate", "price", "cost per unit", "unit cost", "unit rate"],
    "total":       ["total", "total price", "subtotal", "extended price", "line total", "amount"],
}

HEADER_KEYWORDS = {
    "description", "item", "unit", "qty", "quantity",
    "price", "total", "rate", "uom", "particulars",
}

# Column headers that indicate a serial/row-number column — always ignored
SERIAL_HEADERS = {"no", "no.", "s.no", "s.no.", "sr", "sr.", "sn", "s/n",
                  "serial", "seq", "item no", "item no.", "#"}

TOTAL_ROW_PATTERN = re.compile(
    r'^\s*(grand\s+)?total|sub\s*total|lot\s+total|sum\s*$',
    re.IGNORECASE,
)


class ItemsParser:
    # ------------------------------------------------------------------
    # Public
    # ------------------------------------------------------------------

    def parse(self, docs: dict) -> list:
        """
        Returns a flat list of lot dicts:
          {"filename": str, "lot_name": str, "items": pd.DataFrame}
        DataFrame columns: Description, Unit, Quantity, Unit Price, Total
        """
        results = []
        for fname, meta in docs.items():
            ftype = meta.get("file_type", "")
            fpath = meta.get("path", "")
            is_financial = (
                meta.get("label") == "Financial Template"
                or ftype in ("xlsx", "csv")
            )
            if not is_financial:
                continue
            if not fpath or not os.path.exists(fpath):
                continue
            try:
                if ftype == "xlsx":
                    lots = self._parse_xlsx(fpath, fname)
                elif ftype == "csv":
                    lots = self._parse_csv(fpath, fname)
                else:
                    continue
                results.extend(lots)
            except Exception as e:
                results.append({
                    "filename": fname,
                    "lot_name": "⚠ Parse Error",
                    "items": pd.DataFrame(),
                    "error": str(e),
                })
        return results

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def _parse_xlsx(self, fpath: str, fname: str) -> list:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        lots = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            lots.extend(self._extract_lots(rows, fname, sheet_name))

        return lots

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _parse_csv(self, fpath: str, fname: str) -> list:
        df = pd.read_csv(fpath, header=None, encoding="utf-8", on_bad_lines="skip")
        rows = [list(r) for r in df.values.tolist()]
        return self._extract_lots(rows, fname, fname)

    # ------------------------------------------------------------------
    # Core extraction — shared by xlsx and csv
    # ------------------------------------------------------------------

    def _extract_lots(self, rows: list, fname: str, sheet_label: str) -> list:
        lots = []
        current_lot = None
        col_map = {}
        items = []

        for row in rows:
            # Skip fully empty rows
            non_empty = [v for v in row if v is not None and str(v).strip() not in ("", "None", "nan")]
            if not non_empty:
                continue

            lot_name = self._detect_lot_header(row)
            if lot_name:
                if items:
                    lots.append(self._make_lot(fname, current_lot or sheet_label, items))
                current_lot = lot_name
                col_map = {}
                items = []
                continue

            if self._is_column_header(row):
                col_map = self._map_columns(row)
                continue

            if not col_map:
                continue

            item = self._extract_item(row, col_map)
            if item:
                items.append(item)

        if items:
            lots.append(self._make_lot(fname, current_lot or sheet_label, items))

        return lots

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _detect_lot_header(self, row: list) -> str:
        """Return lot name if the row looks like a lot-section header, else None."""
        non_empty = [
            str(v).strip() for v in row
            if v is not None and str(v).strip() not in ("", "None", "nan")
        ]
        if not non_empty:
            return None
        first = non_empty[0]
        # Matches "Lot 1", "Lot A", "LOT 2: Description", etc.
        if LOT_PATTERN.search(first):
            # Passes if ≤2 distinct non-empty cells (merged-cell pattern)
            if len(non_empty) <= 2:
                return first
            # Or if all other non-empty values are identical to the first (repeated merged value)
            if all(v == first for v in non_empty[1:]):
                return first
        return None

    def _is_column_header(self, row: list) -> bool:
        """True if the row contains ≥2 recognisable column-header keywords."""
        hits = 0
        for v in row:
            if v is None:
                continue
            cell = str(v).strip().lower()
            if any(kw in cell for kw in HEADER_KEYWORDS):
                hits += 1
                if hits >= 2:
                    return True
        return False

    def _map_columns(self, header_row: list) -> dict:
        """
        Map canonical field names → column index.
        For each canonical, choose the column with the highest fuzzy score (not just
        the first above threshold). Serial-number columns are excluded from all mappings.
        A column index can only be assigned to one canonical.
        """
        headers = [str(v).strip() if v is not None else "" for v in header_row]

        # Mark which indices are serial-number columns so they are never mapped
        serial_indices = {
            i for i, h in enumerate(headers) if h.lower() in SERIAL_HEADERS
        }

        # Build (canonical → best (score, index)) mapping
        best = {}
        used_indices = set()

        for canonical, candidates in CANONICAL_CANDIDATES.items():
            top_score = 0
            top_idx = None
            for i, h in enumerate(headers):
                if not h or i in serial_indices:
                    continue
                result = rfprocess.extractOne(h.lower(), [c.lower() for c in candidates])
                if result and result[1] > top_score:
                    top_score = result[1]
                    top_idx = i
            if top_idx is not None and top_score >= 70:
                best[canonical] = (top_score, top_idx)

        # Resolve conflicts: if two canonicals want the same column, the higher scorer wins
        col_map = {}
        # Sort by score descending so highest scorer claims the column first
        for canonical, (score, idx) in sorted(best.items(), key=lambda x: -x[1][0]):
            if idx not in used_indices:
                col_map[canonical] = idx
                used_indices.add(idx)

        return col_map

    @staticmethod
    def _is_numeric_only(value: str) -> bool:
        """True if the string is a plain number (row index, not a description)."""
        try:
            float(value.replace(",", ""))
            return True
        except ValueError:
            return False

    def _extract_item(self, row: list, col_map: dict) -> dict:
        """Extract one item dict from a data row using the column map."""
        desc_idx = col_map.get("description")
        if desc_idx is None or desc_idx >= len(row):
            return None
        desc = str(row[desc_idx]).strip() if row[desc_idx] is not None else ""
        if not desc or desc.lower() in ("none", "nan", ""):
            return None
        # Skip rows where description is a plain number (serial / row-index columns)
        if self._is_numeric_only(desc):
            return None
        # Skip summary/total rows
        if TOTAL_ROW_PATTERN.search(desc):
            return None

        item = {"Description": desc}
        for canonical, label in (
            ("unit",       "Unit"),
            ("quantity",   "Quantity"),
            ("unit_price", "Unit Price"),
            ("total",      "Total"),
        ):
            idx = col_map.get(canonical)
            if idx is not None and idx < len(row):
                raw = row[idx]
                item[label] = raw if raw is not None else ""
        return item

    def _make_lot(self, fname: str, lot_name: str, items: list) -> dict:
        return {
            "filename": fname,
            "lot_name": lot_name,
            "items": pd.DataFrame(items),
        }
